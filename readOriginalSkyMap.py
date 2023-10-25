from skymap.SkyMapUtils import read_a_skymap, skymap_standard
from utils import get_args
from env import MultiSatelliteEnv
import openpyxl
from DirectABC import ABC
import healpy as hp
import copy
import numpy as np
import torch
from Methods.train import MLP
from skymap.SkyMapUtils import visualize, visualize_selected_pixel
import time


use_model = False
# 加载模型
input_size = 768+12
hidden_size = 1024
output_size = 6
model = MLP(input_size, hidden_size, output_size)
state_dict = torch.load('Methods/model_mlp.pth', map_location=torch.device('cpu'))
model.load_state_dict(state_dict['model'])
model = model.to(torch.float)
task_mean, task_std = 0.4993, 0.2887
sat_mean, sat_std = 291.4322, 212.5493
action_mean, action_std = 417.2829, 223.8058
reward_mean, reward_std = 93.9235, 71.0114

def state_norm(task,sat):
    norm_task = (task - task.min()) / (task.max()-task.min())
    norm_sat = (sat - sat.min()) / (sat.max()-sat.min())
    norm_task = torch.from_numpy(norm_task)
    norm_sat = torch.from_numpy(norm_sat)
    return torch.cat((norm_task, norm_sat), dim=0).to(torch.float32)
def action_norm_inverse(action):
    return action_mean + action * action_std


args = get_args()
wb = openpyxl.load_workbook('data/eventID.xlsx')
ws = wb['Sheet1']
eventID = []
nrow = ws.max_row
for i in range(nrow - 1):
    cell = ws.cell(row=i + 2, column=1)
    eventID.append(cell.value)

action_batch = []
state_batch = []
reward_batch = []


for event in eventID:
    prob, _, _, _ = read_a_skymap(event=event, random=False)
    m, npix, ra, dec, area = skymap_standard(prob, args.nside_std)
    env = MultiSatelliteEnv(args.num_epoch_steps, m)
    action_num = 0
    state, info = env.reset()
    action_episode = []
    state_episode = []
    reward_episode = []
    terminated = False
    while not terminated:
        # 蜂群算法求解
        time_start = time.time()
        mabc_action, mabc_Val, abc_info = ABC(args, env, state, info)
        time_end = time.time()
        # mabc_ring_action = hp.nest2ring(args.nside_std, mabc_action)
        # 模型求解的动作
        if use_model:
            model.eval()
            with torch.no_grad():
                model_action = model(state_norm(state['task'], state['sat']))
            model_action = torch.round(action_norm_inverse(model_action)).to(torch.int).numpy()
            _, model_reward, _, _ = env.step(model_action, next=False)
        # 可视化动作
        if use_model:
            visualize(m)
            visualize_selected_pixel(m, env.sorted_pixel_indices[model_action], 8, title="model result")
            visualize_selected_pixel(m, env.sorted_pixel_indices[mabc_action], 8, title="abc result")
        # 存储动作状态奖励，进入新状态
        state_episode.append(copy.deepcopy(state))
        action_episode.append(copy.deepcopy(mabc_action))
        state, local_reward, terminated, consflag = env.step(mabc_action, next=True)
        reward_episode.append(copy.deepcopy(local_reward))

        print('--------------------------------------------------------------------------')
        # 输出的动作均是ring排列下的网格索引
        print('正在规划任务{}的第{}个动作:\nABC输出{},约束满足情况{},获得奖励{}'.format(
            event,
            action_num + 1,
            mabc_action,
            consflag.astype(int),
            local_reward
        ), end='')

        if use_model:
            print(',\nmodel输出{},获得奖励{}'.format(
                model_action,
                model_reward
            ), end='')

        print('\n当前运行时间为{}秒'.format(state['t']))
        print('\n搜索用时{}秒'.format(time_end-time_start))
        action_num += 1


    action_batch.append(action_episode)
    state_batch.append(state_episode)
    reward_batch.append(reward_episode)
    action_space = env.sorted_pixel_indices


np.savez('simple_data.npz', data1=action_batch, data2=state_batch, data3=reward_batch, data4=action_space)





