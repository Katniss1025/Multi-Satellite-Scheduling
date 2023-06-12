import os
import pdb

import healpy
import numpy as np
import openpyxl
from healpy.newvisufunc import projview
from matplotlib import pyplot as plt
from pylab import show

import skymap.probUtils as pu
from skymap.transUtils import rotate

from utils import get_args


def data_reinforcement_by_rotate():
    """通过旋转图像实现数据增强
    Returns:
        m_rotated(array): 旋转后的概率
        m_rotated_area_90(float): 旋转后的90%区域面积，单位为平方度
        m_rotated_area_50(float): 旋转后的50%区域面积，单位为平方度
    """
    # Read a fits file
    # root = (os.path.abspath(os.path.join(os.getcwd(), "../")))
    root = os.getcwd()
    wb = openpyxl.load_workbook(root + '/data/eventID.xlsx')
    ws = wb['Sheet1']
    eventID = []
    data = []
    nrow = ws.max_row
    for i in range(nrow - 1):
        cell = ws.cell(row=i + 2, column=1)
        eventID.append(cell.value)
    event = eventID[np.random.randint(len(eventID))]  # 从57条数据中随机选择一条
    try:
        data, h = healpy.read_map(root + '/data/SkyMap/Flat/' + event + '_Flat.fits.gz', h=True)
    except:
        data, h = healpy.read_map(root + '/data/SkyMap/Flat/' + event + '_Flat.fits.gz', h=True)
    # h = dict(h)
    # nside = h['NSIDE']

    # 标准化处理
    args = get_args()
    nside_std = args.nside_std
    m = healpy.ud_grade(data, power=-2, nside_out=nside_std)  # 重采样
    npix = healpy.nside2npix(nside_std)  # 标准healpix下像素总数
    pix_indices = np.arange(npix)
    area = healpy.nside2pixarea(nside=nside_std, degrees=True)

    # 将天球图像转换为坐标
    ra, dec = healpy.pix2ang(nside_std, pix_indices, lonlat=True)

    # Data Reinforcement by rotation
    _ra, _dec, tag = rotate(ra, dec)  # in degree

    # Convert the Equatorial coordinates back to pixel coordinates
    pix_rotated = healpy.ang2pix(nside=nside_std, theta=_ra, phi=_dec, lonlat=True)

    # Recalculate the prob
    m_rotated = m[pix_rotated]
    m_rotated = m_rotated / np.sum(m_rotated)

    # Calculate credible region area
    m_level = pu.cal_credible_level(m)
    m_rotated_level = pu.cal_credible_level(m_rotated)
    m_area_90 = pu.find_credible_region(nside=nside_std, credible_levels=m_level, credible=0.9)
    m_rotated_area_90 = pu.find_credible_region(nside=nside_std, credible_levels=m_rotated_level, credible=0.9)
    m_rotated_area_50 = pu.find_credible_region(nside=nside_std, credible_levels=m_rotated_level, credible=0.5)

    print('图像已生成')

    return m_rotated, m_rotated_area_90, m_rotated_area_50
    # plot SkyMap
    # projview(
    #     m,
    #     coord=["E"],
    #     graticule=True,
    #     cmap=plt.cm.RdYlBu,
    #     cbar=False,
    #     graticule_labels=True,
    #     longitude_grid_spacing=45,
    #     projection_type="mollweide",  # cart, mollweide
    #     # unit="cbar label",
    #     # xlabel="Right Ascension",
    #     # ylabel="Declination",
    #     # cb_orientation="vertical",
    #     # rot=(30,45),
    #     title='Original Map:',
    #     llabel=event,
    #     rlabel='90% area:' + str(round(m_area_90)) + 'deg2'
    # )
    # show()
    # projview(
    #     m_rotated,
    #     coord=["E"],
    #     graticule=True,
    #     cmap=plt.cm.RdYlBu,
    #     cbar=False,
    #     graticule_labels=True,
    #     projection_type="mollweide",  # cart, mollweide
    #     longitude_grid_spacing=45,
    #     title=tag,
    #     llabel=event,
    #     rlabel='90% area:' + str(round(m_rotated_area_90)) + 'deg2'
    #     # unit="cbar label",
    #     # xlabel="Right Ascension",
    #     # ylabel="Declination",
    #     # cb_orientation="vertical",
    #     # rot=(30,45),
    # )
    # show()
    #
    #
