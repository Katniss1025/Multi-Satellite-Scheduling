import numpy as np
import os
import openpyxl
import healpy
from healpy.newvisufunc import projview
from matplotlib import pyplot as plt
from pylab import show


def read_a_skymap(event=None, random=True):
    # root = (os.path.abspath(os.path.join(os.getcwd(), "../")))
    root = os.getcwd()
    if random == True:
        wb = openpyxl.load_workbook(root + '/data/eventID.xlsx')
        ws = wb['Sheet1']
        eventID = []
        nrow = ws.max_row
        for i in range(nrow - 1):
            cell = ws.cell(row=i + 2, column=1)
            eventID.append(cell.value)
        event = eventID[np.random.randint(len(eventID))]  # 从57条数据中随机选择一条
    data_ring, h = healpy.read_map(root + '/data/SkyMap/Flat/'+event+'_Flat.fits.gz', h=True, field=0)
    data_nest = healpy.read_map(root + '/data/SkyMap/Flat/'+event+'_Flat.fits.gz', nest=True)
    return data_ring, h, event, data_nest


def skymap_standard(prob, nside_std):
    m = healpy.ud_grade(prob, power=-2, nside_out=nside_std)  # 重采样
    npix = healpy.nside2npix(nside_std)  # 标准healpix下像素总数
    pix_indices = np.arange(npix)
    area = healpy.nside2pixarea(nside=nside_std, degrees=True)
    ra, dec = healpy.pix2ang(nside_std, pix_indices, lonlat=True)
    return m, npix, ra, dec, area


def visualize(prob, title=None):
    projview(
        prob,
        coord=["E"],
        graticule=True,
        cmap=plt.cm.RdYlBu,
        cbar=False,
        graticule_labels=True,
        longitude_grid_spacing=45,
        projection_type="mollweide",  # cart, mollweide
        title=title,
    )
    show()


def interpolate_sky_map(m, nside, image=False):
    # 将m从ring映射到nest顺序
    import healpy
    m_nest = healpy.reorder(m, r2n=True)

    # 定义目标二维图像的网格坐标
    import numpy as np
    from astropy import units as u
    ra = np.linspace(360., 0., 361) * u.deg
    dec = np.linspace(-90., 90., 181) * u.deg
    ra_grid, dec_grid = np.meshgrid(ra, dec)

    # Set up Astropy coordinate objects
    from astropy.coordinates import SkyCoord
    from astropy.coordinates import Galactic
    coords = SkyCoord(ra_grid.ravel(), dec_grid.ravel(), frame=Galactic())  # .ravel()把所有的拉成一行

    from astropy_healpix import HEALPix
    hp = HEALPix(nside=nside, order='NESTED', frame=Galactic())
    pmap = hp.interpolate_bilinear_skycoord(coords, m_nest)
    pmap = pmap.reshape((181, 361))

    # plot
    import matplotlib.pyplot as plt
    if image:
        plt.figure(figsize=(10, 5))
        im = plt.imshow(pmap, extent=[360, 0, -90, 90], cmap=plt.cm.RdYlBu, origin='lower', aspect='auto')
        plt.colorbar(im)
        plt.xlabel('Right ascension')
        plt.ylabel('Declination')
        plt.grid()
        plt.show()

    return pmap


def visualize_selected_pixel(m, pixel_indices, nside):
    import healpy as hp
    import numpy as np
    import matplotlib.pyplot as plt

    npix = hp.nside2npix(nside)  # 总像素数

    # 将需要可视化的像素赋予特定的值，其他像素的值设为 NaN
    mask = np.zeros(npix)
    mask[pixel_indices] = m[pixel_indices]
    mask[~np.in1d(np.arange(npix), pixel_indices)] = np.nan

    # 可视化
    hp.mollview(mask, title="Selected Pixels")
    plt.show()



